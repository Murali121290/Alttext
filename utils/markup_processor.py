"""
markup_processor.py — Utility for the PDF Markup Tool.

Crops user-drawn regions from a PDF page and sends each crop to Gemini
to generate WCAG-compliant alt text. Also writes results to Excel.

NOTE: This module is purely additive and does not touch the existing
batch processing pipeline in worker_tasks.py.
"""

import os
import io
import re
import json
import base64
import logging
import tempfile
import threading

import fitz  # PyMuPDF
from PIL import Image
import google.genai as genai
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.utils import get_column_letter
from dotenv import load_dotenv

load_dotenv()
logger = logging.getLogger(__name__)

GEMINI_API_KEY = os.getenv("GEMINI_API_KEY")
MODEL_NAME = "gemini-2.5-pro"

# Thread-safe lazy client initialisation
_client = None
_client_lock = threading.Lock()


def _get_client():
    global _client
    if _client is None:
        with _client_lock:
            if _client is None:
                if not GEMINI_API_KEY:
                    raise ValueError("GEMINI_API_KEY not set in environment")
                
                # Check for SSL_VERIFY configuration to bypass SSL certificate checks if needed
                import ssl
                ssl_verify_str = os.getenv("SSL_VERIFY", "True").lower()
                ssl_verify = ssl_verify_str not in ("false", "0", "no", "off")
                
                http_opts = {"timeout": 120000}
                if not ssl_verify:
                    unverified_ssl_context = ssl.create_default_context()
                    unverified_ssl_context.check_hostname = False
                    unverified_ssl_context.verify_mode = ssl.CERT_NONE
                    http_opts["client_args"] = {"verify": unverified_ssl_context}
                    http_opts["async_client_args"] = {"verify": unverified_ssl_context}
                    logger.info("Markup client: Outbound Gemini API SSL verification disabled.")

                _client = genai.Client(
                    api_key=GEMINI_API_KEY,
                    http_options=http_opts
                )
    return _client


MARKUP_SYSTEM_PROMPT = (
    "You are an expert alt text authoring assistant for WCAG 2.2 compliance.\n\n"
    "The user has cropped a specific region from a PDF page. The crop may contain "
    "a figure/image, a table, an equation, a text block, or a mix of these.\n\n"
    "Analyze the crop carefully and return ONLY valid JSON — no markdown fences, "
    "no explanation outside the JSON object:\n\n"
    "{\n"
    '  "content_type": "image|table|equation|text|mixed",\n'
    '  "short_alt": "5-25 word alt text",\n'
    '  "long_alt": "Full WCAG-compliant alt text, 25-300 words"\n'
    "}\n\n"
    "Rules:\n"
    "- No leading articles (A, An, The) at the start of any alt text field\n"
    "- Never use 'image of', 'picture of', 'shows', 'depicts', 'illustrates'\n"
    "- For equations: describe using plain English mathematical language "
    "(e.g., 'integral from 0 to infinity of e to the negative x equals 1')\n"
    "- For tables: describe structure, column headers, and key data values\n"
    "- For text blocks: transcribe verbatim if short; summarize clearly if long\n"
    "- For figures/images: describe content and purpose, not aesthetics\n"
    "- Use present tense, active voice, American English spelling\n"
    "- Every sentence must contain a verb\n"
    "- Hard cap: long_alt must not exceed 300 words\n"
)


def crop_region_to_png(pdf_path: str, page_num: int,
                        x0_pct: float, y0_pct: float,
                        x1_pct: float, y1_pct: float) -> bytes:
    """
    Crop a percentage-based bounding box from a PDF page and return PNG bytes.
    Coordinates are fractions of page width/height (0.0 – 1.0).
    Rendered at 2x scale for quality.
    """
    for name, val in [("x0_pct", x0_pct), ("y0_pct", y0_pct), ("x1_pct", x1_pct), ("y1_pct", y1_pct)]:
        if not (0.0 <= val <= 1.0):
            raise ValueError(f"Coordinate {name}={val} is out of range [0.0, 1.0]")
    if x0_pct >= x1_pct or y0_pct >= y1_pct:
        raise ValueError(f"Invalid crop rect: ({x0_pct},{y0_pct}) → ({x1_pct},{y1_pct})")

    doc = fitz.open(pdf_path)
    try:
        page = doc[page_num]
        pw = page.rect.width
        ph = page.rect.height
        clip = fitz.Rect(
            x0_pct * pw,
            y0_pct * ph,
            x1_pct * pw,
            y1_pct * ph,
        )
        mat = fitz.Matrix(2, 2)
        pix = page.get_pixmap(matrix=mat, clip=clip, alpha=False, annots=False)
        return pix.tobytes("png")
    finally:
        doc.close()


MARKUP_REFINE_PROMPT_TMPL = (
    "You previously generated the following alt text for a cropped PDF region:\n\n"
    "Short alt: {short_alt}\n"
    "Long alt: {long_alt}\n\n"
    "The user wants you to refine it with this instruction: \"{prompt}\"\n\n"
    "Return ONLY valid JSON — no markdown fences, no explanation:\n"
    '{{"short_alt": "...", "long_alt": "..."}}\n\n'
    "Apply all original WCAG rules: no leading articles (A/An/The), no banned phrases "
    "('image of', 'shows', 'depicts'), present tense, active voice, American English, "
    "hard cap of 300 words for long_alt."
)


def refine_markup_region(png_bytes: bytes, previous_short: str,
                          previous_long: str, prompt: str) -> dict:
    """
    Refine existing alt text for a region based on a user instruction.
    Re-sends the crop image so Gemini can reference the visual while revising.
    Returns dict with keys: short_alt, long_alt.
    """
    client = _get_client()

    image = Image.open(io.BytesIO(png_bytes))
    if image.mode != "RGB":
        image = image.convert("RGB")
    if image.width > 2048 or image.height > 2048:
        image.thumbnail((2048, 2048), Image.LANCZOS)

    refine_prompt = MARKUP_REFINE_PROMPT_TMPL.format(
        short_alt=previous_short,
        long_alt=previous_long,
        prompt=prompt,
    )

    MAX_RETRIES = 5
    INITIAL_WAIT = 1
    import time
    for attempt in range(MAX_RETRIES):
        try:
            response = client.models.generate_content(
                model=MODEL_NAME,
                contents=[refine_prompt, image],
                config={"temperature": 0.3, "top_p": 0.9},
            )
            break
        except Exception as e:
            error_msg = str(e)
            is_transient_error = (
                "429" in error_msg or 
                "502" in error_msg or 
                "500" in error_msg or 
                "503" in error_msg or 
                "quota" in error_msg.lower() or
                "timed out" in error_msg.lower() or
                "timeout" in error_msg.lower() or
                "connection" in error_msg.lower()
            )
            if is_transient_error and attempt < MAX_RETRIES - 1:
                sleep_time = INITIAL_WAIT * (2 ** (attempt + 1))
                logger.warning(f"Markup Refine Gemini transient error: {e}. Retrying in {sleep_time}s (Attempt {attempt+1}/{MAX_RETRIES})...")
                time.sleep(sleep_time)
                continue
            raise

    raw = (response.text or "").strip()
    code_block = re.search(r"```(?:json)?\s*(.*?)```", raw, re.DOTALL)
    if code_block:
        json_str = code_block.group(1).strip()
    else:
        m = re.search(r"\{.*\}", raw, re.DOTALL)
        json_str = m.group() if m else raw

    try:
        return json.loads(json_str)
    except json.JSONDecodeError:
        logger.warning(f"Markup refine: response not valid JSON: {raw[:200]}")
        return {"short_alt": previous_short, "long_alt": raw}


def call_gemini_for_region(png_bytes: bytes) -> dict:
    """
    Send a cropped region PNG to Gemini 2.5 Pro and return a parsed dict
    with keys: content_type, short_alt, long_alt.
    """
    client = _get_client()

    image = Image.open(io.BytesIO(png_bytes))
    if image.mode != "RGB":
        image = image.convert("RGB")
    # Cap size to avoid excessive token usage
    if image.width > 2048 or image.height > 2048:
        image.thumbnail((2048, 2048), Image.LANCZOS)

    MAX_RETRIES = 5
    INITIAL_WAIT = 1
    import time
    for attempt in range(MAX_RETRIES):
        try:
            response = client.models.generate_content(
                model=MODEL_NAME,
                contents=[MARKUP_SYSTEM_PROMPT, image],
                config={"temperature": 0.1, "top_p": 0.9},
            )
            break
        except Exception as e:
            error_msg = str(e)
            is_transient_error = (
                "429" in error_msg or 
                "502" in error_msg or 
                "500" in error_msg or 
                "503" in error_msg or 
                "quota" in error_msg.lower() or
                "timed out" in error_msg.lower() or
                "timeout" in error_msg.lower() or
                "connection" in error_msg.lower()
            )
            if is_transient_error and attempt < MAX_RETRIES - 1:
                sleep_time = INITIAL_WAIT * (2 ** (attempt + 1))
                logger.warning(f"Markup Gemini transient error: {e}. Retrying in {sleep_time}s (Attempt {attempt+1}/{MAX_RETRIES})...")
                time.sleep(sleep_time)
                continue
            raise

    raw = (response.text or "").strip()

    # Try code-fence first, then bare JSON object
    code_block = re.search(r"```(?:json)?\s*(.*?)```", raw, re.DOTALL)
    if code_block:
        json_str = code_block.group(1).strip()
    else:
        m = re.search(r"\{.*\}", raw, re.DOTALL)
        json_str = m.group() if m else raw

    try:
        return json.loads(json_str)
    except json.JSONDecodeError:
        logger.warning(f"Markup: Gemini response not valid JSON: {raw[:200]}")
        return {
            "content_type": "unknown",
            "short_alt": "",
            "long_alt": raw,
        }


def process_markup_regions(pdf_path: str, regions: list) -> list:
    """
    Process a list of user-drawn regions through Gemini.

    Each region dict must have:
        id, page (0-based int), x0_pct, y0_pct, x1_pct, y1_pct (floats 0-1),
        label (str, may be empty)

    Returns the same list with added keys:
        content_type, short_alt, long_alt, crop_png_b64
    """
    results = []
    for region in regions:
        try:
            png_bytes = crop_region_to_png(
                pdf_path,
                int(region["page"]),
                float(region["x0_pct"]),
                float(region["y0_pct"]),
                float(region["x1_pct"]),
                float(region["y1_pct"]),
            )
            gemini_result = call_gemini_for_region(png_bytes)
            results.append({
                **region,
                "content_type": gemini_result.get("content_type", "unknown"),
                "short_alt": gemini_result.get("short_alt", ""),
                "long_alt": gemini_result.get("long_alt", ""),
                "crop_png_b64": base64.b64encode(png_bytes).decode(),
            })
        except Exception as e:
            logger.error(f"Markup: error processing region {region.get('id')}: {e}")
            results.append({
                **region,
                "content_type": "error",
                "short_alt": "",
                "long_alt": f"Processing error: {e}",
                "crop_png_b64": "",
            })
    return results


def _apply_json_rules(text: str) -> str:
    """Mirror of worker_tasks.apply_json_rules_to_alt_text — strips banned phrases."""
    if not text:
        return ""
    try:
        rules_path = os.path.join(os.path.dirname(__file__), "alt_text_rules.json")
        if os.path.exists(rules_path):
            with open(rules_path, "r", encoding="utf-8") as f:
                rules_data = json.load(f)
            for rule_data in rules_data.get("alt_text_validation_rules", {}).values():
                for phrase in rule_data.get("words", []):
                    pattern = r"(?<!\w)" + re.escape(phrase) + r"(?!\w)\s*"
                    text = re.sub(pattern, "", text, flags=re.IGNORECASE)
            text = text.strip()
            if text and text[0].islower():
                text = text[0].upper() + text[1:]
    except Exception as e:
        logger.error(f"Markup: error applying alt text rules: {e}")
    return text


def _clean_alt_text(text: str) -> str:
    """Mirror of worker_tasks.clean_alt_text — sanitises for Excel."""
    if not text:
        return ""
    text = re.sub(r"[\x00-\x08\x0b-\x0c\x0e-\x1f]", "", text)
    text = text.lstrip("=+-@").strip()
    text = re.sub(
        r"[^.!?]*\b(copyright|©|photo credit|photo by|courtesy of|attributed to|credit:)[^.!?]*[.!?]?",
        "", text, flags=re.IGNORECASE,
    ).strip()
    text = re.sub(r"\b(Figure|Fig\.?)\s*[\d.]+[:\s]+", "", text, flags=re.IGNORECASE)
    text = re.sub(r"^(?:a|an|the)\s+", "", text, flags=re.IGNORECASE)
    return text


def write_markup_excel(results: list, output_path: str, pdf_filename: str = "") -> None:
    """
    Write markup results to an Excel workbook matching the batch pipeline format:
    File name | Figure number | Page number | Image | Short alt text |
    Long alt text | Word Count | Category | Context Type | Domain
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Alt Text"

    headers = [
        "File name", "Figure number", "Page number", "Image",
        "Short alt text", "Long alt text", "Word Count", "Category",
        "Context Type", "Domain",
    ]
    ws.append(headers)

    # Sort by page then by region id
    sorted_results = sorted(results, key=lambda r: (r.get("page", 0), r.get("id", 0)))
    _tmp_paths = []

    base_name = os.path.splitext(pdf_filename)[0] if pdf_filename else "markup"

    for row_idx, result in enumerate(sorted_results, 2):
        raw_short = result.get("short_alt", "")
        raw_long = result.get("long_alt", "")

        final_short = _clean_alt_text(_apply_json_rules(raw_short))
        final_long = _clean_alt_text(_apply_json_rules(raw_long))

        word_count = len(final_long.split()) if final_long else 0
        if word_count < 25:
            category = "Simple"
        elif word_count < 150:
            category = "Moderate"
        else:
            category = "Complex"

        label = result.get("label", "").strip()
        page_1based = int(result.get("page", 0)) + 1
        figure_number = label if label else f"Figure {row_idx - 1}"

        row = [
            pdf_filename,
            figure_number,
            page_1based,
            "",          # Image column — thumbnail embedded below
            final_short,
            final_long,
            word_count,
            category,
            result.get("context_type", "General"),
            result.get("domain", "General"),
        ]
        ws.append(row)

        # Embed crop thumbnail in column D
        crop_b64 = result.get("crop_png_b64", "")
        if crop_b64:
            try:
                img_bytes = base64.b64decode(crop_b64)
                pil_img = Image.open(io.BytesIO(img_bytes))
                if pil_img.mode != "RGB":
                    pil_img = pil_img.convert("RGB")

                max_size = 200
                ratio = min(max_size / pil_img.width, max_size / pil_img.height)
                if ratio < 1:
                    pil_img = pil_img.resize(
                        (int(pil_img.width * ratio), int(pil_img.height * ratio)),
                        Image.LANCZOS,
                    )

                # openpyxl requires a real file path or seekable BytesIO
                with tempfile.NamedTemporaryFile(suffix=".png", delete=False) as tmp:
                    tmp_path = tmp.name
                pil_img.save(tmp_path, format="PNG")

                xl_img = OpenpyxlImage(tmp_path)
                ws.add_image(xl_img, f"D{row_idx}")
                ws.row_dimensions[row_idx].height = (xl_img.height * 0.75) + 10
                current_col_width = ws.column_dimensions["D"].width or 10
                needed_width = (xl_img.width / 7) + 2
                if needed_width > current_col_width:
                    ws.column_dimensions["D"].width = needed_width
                _tmp_paths.append(tmp_path)
            except Exception as e:
                logger.warning(f"Markup Excel: could not embed thumbnail row {row_idx}: {e}")

    try:
        wb.save(output_path)
    finally:
        for _tmp in _tmp_paths:
            try:
                os.unlink(_tmp)
            except Exception:
                pass


def _embed_crop_image(ws, crop_b64: str, excel_row: int, col_letter: str = "D") -> str | None:
    """Decode a base64 crop PNG and embed it in the given cell. Returns temp path or None."""
    if not crop_b64:
        return None
    try:
        img_bytes = base64.b64decode(crop_b64)
        pil_img = Image.open(io.BytesIO(img_bytes))
        if pil_img.mode != "RGB":
            pil_img = pil_img.convert("RGB")
        max_size = 200
        ratio = min(max_size / pil_img.width, max_size / pil_img.height)
        if ratio < 1:
            pil_img = pil_img.resize(
                (int(pil_img.width * ratio), int(pil_img.height * ratio)),
                Image.LANCZOS,
            )
        with tempfile.NamedTemporaryFile(suffix=".png", delete=False) as tmp:
            tmp_path = tmp.name
        pil_img.save(tmp_path, format="PNG")
        xl_img = OpenpyxlImage(tmp_path)
        ws.add_image(xl_img, f"{col_letter}{excel_row}")
        return tmp_path
    except Exception:
        return None


def update_review_excel(results: list, excel_path: str) -> None:
    """
    Rebuild the Excel file from the frontend's complete region list.

    The frontend sends every current region (including crop_png_b64 loaded
    from the original file), so we can do a clean full-rebuild instead of
    a fragile merge.  This eliminates duplicate rows caused by delete+redraw.

    Column layout:
      A: File name  B: Figure number  C: Page number  D: Image (thumbnail)
      E: Short alt  F: Long alt  G: Word Count  H: Category  I: Context Type  J: Domain
    """
    from openpyxl import load_workbook

    # Read pdf_filename from the existing file before overwriting it
    pdf_filename = ""
    try:
        wb_old = load_workbook(excel_path, read_only=True, data_only=True)
        ws_old = wb_old.active
        for row in ws_old.iter_rows(min_row=2, max_row=2, values_only=True):
            pdf_filename = str(row[0]) if row[0] else ""
            break
        wb_old.close()
    except Exception:
        pass

    # Sort by page then by region id to preserve reading order
    sorted_results = sorted(results, key=lambda r: (r.get("page", 0), r.get("id", 0)))

    wb = Workbook()
    ws = wb.active
    ws.title = "Alt Text"

    headers = [
        "File name", "Figure number", "Page number", "Image",
        "Short alt text", "Long alt text", "Word Count", "Category",
        "Context Type", "Domain",
    ]
    ws.append(headers)

    _tmp_paths = []

    for row_idx, result in enumerate(sorted_results, 2):
        raw_short = result.get("short_alt", "")
        raw_long  = result.get("long_alt", "")

        final_short = _clean_alt_text(_apply_json_rules(raw_short))
        final_long  = _clean_alt_text(_apply_json_rules(raw_long))

        word_count = len(final_long.split()) if final_long else 0
        if word_count < 25:
            category = "Simple"
        elif word_count < 150:
            category = "Moderate"
        else:
            category = "Complex"

        label       = result.get("label", "").strip() or f"Figure {row_idx - 1}"
        page_1based = int(result.get("page", 0)) + 1

        ws.append([
            pdf_filename,
            label,
            page_1based,
            "",          # D: thumbnail embedded below
            final_short,
            final_long,
            word_count,
            category,
            result.get("content_type", "General"),
            result.get("domain", "General"),
        ])

        crop_b64 = result.get("crop_png_b64") or ""
        if crop_b64:
            try:
                img_bytes = base64.b64decode(crop_b64)
                pil_img = Image.open(io.BytesIO(img_bytes))
                if pil_img.mode != "RGB":
                    pil_img = pil_img.convert("RGB")
                max_size = 200
                ratio = min(max_size / pil_img.width, max_size / pil_img.height)
                if ratio < 1:
                    pil_img = pil_img.resize(
                        (int(pil_img.width * ratio), int(pil_img.height * ratio)),
                        Image.LANCZOS,
                    )
                with tempfile.NamedTemporaryFile(suffix=".png", delete=False) as tmp:
                    tmp_path = tmp.name
                pil_img.save(tmp_path, format="PNG")
                xl_img = OpenpyxlImage(tmp_path)
                ws.add_image(xl_img, f"D{row_idx}")
                ws.row_dimensions[row_idx].height = (xl_img.height * 0.75) + 10
                needed_width = (xl_img.width / 7) + 2
                if (ws.column_dimensions["D"].width or 10) < needed_width:
                    ws.column_dimensions["D"].width = needed_width
                _tmp_paths.append(tmp_path)
            except Exception as e:
                logger.warning(f"Review Excel: could not embed thumbnail row {row_idx}: {e}")

    try:
        wb.save(excel_path)
    finally:
        for _tmp in _tmp_paths:
            try:
                os.unlink(_tmp)
            except Exception:
                pass
