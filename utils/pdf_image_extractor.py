import fitz  # PyMuPDF
import os
import logging
from PIL import Image
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage

logger = logging.getLogger(__name__)


def extract_images_to_excel(pdf_path, output_folder="extracted_images", excel_filename="image_report.xlsx"):
    """
    Extracts all embedded images from a PDF, merging tiled/split images back
    into single images, and generates an Excel report with thumbnails.

    Args:
        pdf_path (str): Path to the input PDF file.
        output_folder (str): Directory where extracted images will be saved.
        excel_filename (str): Name for the final Excel report.

    Returns:
        str: Path to the generated Excel file, or None if failed.
    """
    if not os.path.exists(pdf_path):
        logger.error(f"Input PDF not found: {pdf_path}")
        return None

    os.makedirs(output_folder, exist_ok=True)
    excel_file = os.path.join(output_folder, excel_filename)

    try:
        doc = fitz.open(pdf_path)
    except Exception as e:
        logger.error(f"Failed to open PDF {pdf_path}: {e}")
        return None

    wb = Workbook()
    ws = wb.active
    ws.title = "Images"
    ws.append(["Page No", "Image Name", "Image"])

    row = 2
    img_counter = 1
    total_extracted = 0

    for page_num in range(len(doc)):
        try:
            page = doc.load_page(page_num)

            # --- Strategy: Render entire page regions containing images ---
            # Get image bounding boxes on the page
            image_list = page.get_images(full=True)

            if not image_list:
                continue

            logger.info(f"Page {page_num+1} - found {len(image_list)} raw image xrefs")

            # Collect all image bounding boxes on this page
            image_rects = []
            if image_list:
                for img in image_list:
                    xref = img[0]
                    # Get all locations (bbox) where this xref appears on the page
                    rects = page.get_image_rects(xref)
                    for rect in rects:
                        image_rects.append(rect)
            else:
                # If no raster images, check if it's a pure vector graphics figure
                paths = page.get_drawings()
                if paths:
                    logger.info(f"Page {page_num+1} - found {len(paths)} vector paths (no raster images)")
                    for p in paths:
                        r = p["rect"]
                        # Ignore tiny 1x1 vector noise/lines or massive full-page backgrounds
                        if r.width > 5 and r.height > 5 and (r.width < page.rect.width * 0.95 or r.height < page.rect.height * 0.95):
                            image_rects.append(fitz.Rect(r))

            if not image_rects:
                continue

            # Merge overlapping/adjacent rects to detect tiled groups
            merged_rects = merge_rects(image_rects)

            logger.info(f"Page {page_num+1} - merged into {len(merged_rects)} logical image(s)")

            for rect in merged_rects:
                image_name = f"page{page_num+1}_img{img_counter}.jpeg"
                image_path = os.path.join(output_folder, image_name)

                try:
                    # Render just this region at high resolution (2x zoom)
                    clip = fitz.Rect(rect)
                    mat = fitz.Matrix(2, 2)
                    pix = page.get_pixmap(matrix=mat, clip=clip)
                    pil_img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
                    pil_img.save(image_path, "JPEG", quality=95)
                except Exception as e:
                    logger.warning(f"Failed rendering region on page {page_num+1}: {e}")
                    continue

                # Write to Excel
                ws.cell(row=row, column=1, value=page_num + 1)
                ws.cell(row=row, column=2, value=image_name)

                try:
                    xl_img = XLImage(image_path)
                    xl_img.width = 120
                    xl_img.height = 90
                    ws.add_image(xl_img, f"C{row}")
                    ws.row_dimensions[row].height = 70
                except Exception as e:
                    logger.warning(f"Failed embedding {image_name} into Excel: {e}")

                row += 1
                img_counter += 1
                total_extracted += 1

        except Exception as e:
            logger.error(f"Error processing page {page_num+1}: {e}")

    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 25

    try:
        wb.save(excel_file)
        doc.close()
        logger.info(f"Extraction complete. {total_extracted} image(s) saved.")
        logger.info(f"Output folder: {output_folder}")
        logger.info(f"Excel report: {excel_file}")
        return excel_file
    except Exception as e:
        logger.error(f"Failed to save Excel report: {e}")
        doc.close()
        return None


def merge_rects(rects, tolerance=5):
    """
    Merges a list of fitz.Rect objects that overlap or are within `tolerance`
    pixels of each other into unified bounding boxes.

    This handles PDFs that split one logical image into multiple tiles.
    """
    if not rects:
        return []

    # Convert to list of [x0, y0, x1, y1] for easy manipulation
    boxes = [[r.x0, r.y0, r.x1, r.y1] for r in rects]
    merged = True

    while merged:
        merged = False
        result = []
        used = [False] * len(boxes)

        for i in range(len(boxes)):
            if used[i]:
                continue
            current = boxes[i][:]
            for j in range(i + 1, len(boxes)):
                if used[j]:
                    continue
                # Check if boxes overlap or are within tolerance
                if boxes_overlap(current, boxes[j], tolerance):
                    # Expand current to encompass both
                    current[0] = min(current[0], boxes[j][0])
                    current[1] = min(current[1], boxes[j][1])
                    current[2] = max(current[2], boxes[j][2])
                    current[3] = max(current[3], boxes[j][3])
                    used[j] = True
                    merged = True
            result.append(current)
            used[i] = True
        boxes = result

    return [fitz.Rect(b[0], b[1], b[2], b[3]) for b in boxes]


def boxes_overlap(a, b, tolerance=5):
    """Returns True if two [x0,y0,x1,y1] boxes overlap or are within tolerance."""
    return not (
        a[2] + tolerance < b[0] or
        b[2] + tolerance < a[0] or
        a[3] + tolerance < b[1] or
        b[3] + tolerance < a[1]
    )


if __name__ == '__main__':
    logging.basicConfig(level=logging.INFO, format='%(levelname)s: %(message)s')

    import sys
    target = sys.argv[1] if len(sys.argv) > 1 else 'input.pdf'

    if not os.path.exists(target):
        print(f"Usage: python pdf_image_extractor.py <path_to_pdf>")
        print(f"File '{target}' not found.")
    else:
        extract_images_to_excel(target)
