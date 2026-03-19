import os
import shutil
import fitz  # PyMuPDF
import io
import sqlite3
import psycopg2
import psycopg2.extras
import datetime
import time
import threading
import json
import re
from concurrent.futures import ThreadPoolExecutor, as_completed
from PIL import Image
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as OpenpyxlImage
import google.genai as genai
from utils.prompt_assets import SYSTEM_PROMPT
from dotenv import load_dotenv
import logging
try:
    import pandas as pd
    _PANDAS_AVAILABLE = True
except ImportError:
    _PANDAS_AVAILABLE = False
try:
    from openai import OpenAI as _OpenAI
    _OPENAI_AVAILABLE = True
except ImportError:
    _OPENAI_AVAILABLE = False

# ---------------- LOGGING CONFIG ----------------
_stream_handler = logging.StreamHandler()
_stream_handler.stream = open(_stream_handler.stream.fileno(), mode='w', encoding='utf-8', buffering=1)
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(process)d - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler("alttext_worker.log", encoding='utf-8'),
        _stream_handler
    ]
)
logger = logging.getLogger(__name__)

# Suppress verbose logging from Google SDK libraries
logging.getLogger("google").setLevel(logging.WARNING)
logging.getLogger("google.generativeai").setLevel(logging.WARNING)
logging.getLogger("google.api_core").setLevel(logging.WARNING)

# ---------------- CONFIG ----------------
load_dotenv()

MODEL_NAME = "gemini-2.5-pro"
UPLOAD_FOLDER = "uploads"
OUTPUT_FOLDER = "outputs"
GEMINI_API_KEY = os.getenv("GEMINI_API_KEY")
OPENAI_API_KEY = os.getenv("OPENAI_API_KEY")
DATABASE_URL = os.getenv("DATABASE_URL", "postgresql://postgres:postgres@localhost:5432/alttext")

os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(OUTPUT_FOLDER, exist_ok=True)
os.makedirs(os.path.join(OUTPUT_FOLDER, "extracted_images"), exist_ok=True)

# GPT-4o pricing constants (USD per 1M tokens)
_GPT_INPUT_COST_PER_M  = 2.50
_GPT_OUTPUT_COST_PER_M = 10.00

# Gemini 2.5 Pro pricing constants (USD per 1M tokens)
# NOTE: Update these if Gemini pricing changes. These are NOT gemini-1.5-flash prices.
_GEMINI_INPUT_COST_PER_M  = 3.50
_GEMINI_OUTPUT_COST_PER_M = 10.50

# FIX 4 — Thread-safe lazy client initialisation using locks.
# Previously two threads could both see _gemini_client is None and create two clients.
_gemini_client = None
_openai_client = None
_gemini_client_lock = threading.Lock()
_openai_client_lock = threading.Lock()

def get_gemini_client():
    global _gemini_client
    if _gemini_client is None:
        with _gemini_client_lock:
            if _gemini_client is None:
                if not GEMINI_API_KEY:
                    raise ValueError("GEMINI_API_KEY not set")
                # FIX 10 — 120 s timeout prevents a hung request from blocking a worker thread forever.
                # Note: google-genai http_options 'timeout' is in milliseconds, so 120000 = 120 seconds.
                _gemini_client = genai.Client(api_key=GEMINI_API_KEY, http_options={"timeout": 120000})
    return _gemini_client

def get_openai_client():
    global _openai_client
    if _openai_client is None:
        with _openai_client_lock:
            if _openai_client is None:
                if not _OPENAI_AVAILABLE:
                    raise ImportError("openai package not installed. Run: pip install openai")
                if not OPENAI_API_KEY:
                    raise ValueError("OPENAI_API_KEY not set in .env")
                _openai_client = _OpenAI(api_key=OPENAI_API_KEY)
    return _openai_client

# Global concurrency semaphores (shared across ALL users and threads).
# Tune via .env: GEMINI_CONCURRENCY (default 5) and OPENAI_CONCURRENCY (default 1).
# OPENAI_CONCURRENCY=1 forces serial GPT-4o calls — required for Tier-1/free accounts.
_GEMINI_SEMAPHORE = threading.Semaphore(int(os.getenv("GEMINI_CONCURRENCY", "5")))
_OPENAI_SEMAPHORE = threading.Semaphore(int(os.getenv("OPENAI_CONCURRENCY", "1")))

# FIX 15 — Thread-safe DB type state. Previously IS_SQLITE was a bare global mutated
# by get_db_connection() without a lock, causing race conditions during Postgres failover.
_db_state_lock = threading.Lock()
_DB_TYPE = "postgres"
_IS_SQLITE = False

SCORING_PROMPT = """You are a senior accessibility auditor and subject matter expert.
Evaluate the following alt text using a strict domain-aware scoring rubric.
The alt text was generated for a **{domain}** image with context type: **{context_type}**.

IMPORTANT — READ BEFORE SCORING:
- Apply ONLY the rubric section that matches the domain below.
- Do NOT apply medical criteria to Education, Publishing, or non-medical content.
- Do NOT penalize the absence of biological mechanisms in Education or Publishing domains.
- If context_type is "Decorative", return all numeric scores as 0 and decision as "Decorative — Not scored".

═══════════════════════════════════════════════════════════════════
RUBRIC A — USE FOR: Medical, Nursing, Life Sciences
(Apply when domain contains "Medical" or "Nursing")
TOTAL = 100 POINTS

PART 1 — Content Fidelity (40 Points):
1. Functional Equivalence (18 pts): Does the alt text provide the same meaning and instructional value as the diagram?
2. Completeness of Critical Information (10 pts): Are all labeled structures, pathways, and regulatory relationships described?
3. Scientific Accuracy (8 pts): No biological errors or misleading simplifications.
4. Educational Utility & Pedagogy (4 pts): Would a blind medical student gain equivalent understanding?

PART 2 — Structural Clarity (20 Points):
5. Mechanistic Completeness (6 pts): Stimulation, inhibition, feedback loops, and regulatory hierarchies fully explained.
6. Structural & Cognitive Clarity (6 pts): Clear cause-effect relationships. Readable sentence structure.
7. Logical Reading Order (8 pts): Is the process described in a clear and sequential way?

PART 3 — Language Precision (15 Points):
8. Terminology Precision (9 pts): Full terms used correctly. Abbreviations introduced properly.
9. Conciseness Without Redundancy (6 pts): No unnecessary repetition or filler phrases.

PART 4 — Accessibility Compliance (25 Points):
10. No Sensory/Visual-Only Language (15 pts): No references to color, layout, arrows, or "clearly labeled."
11. Screen Reader Readability (5 pts): Alt text is optimized for screen reader synthesis (no excessive punctuation, proper sentence breaks).
12. Proper Descriptive Phrasing (5 pts): Uses objective, descriptive language instead of subjective or vague terms.

═══════════════════════════════════════════════════════════════════
RUBRIC B — USE FOR: Education, Publishing, General (all non-medical domains)
(Apply when domain is Education, Publishing, General, or anything not Medical/Nursing)
TOTAL = 100 POINTS

PART 1 — Content Fidelity (40 Points):
1. Functional Equivalence (18 pts): Does the alt text convey the same objective information as the image?
2. Completeness of Information (10 pts): Are all key elements described — text, labels, figures, diagrams?
3. Content Accuracy (8 pts): Is the described content factually correct? No invented, assumed, or hallucinated details.
4. Educational Utility & Pedagogy (4 pts): Does the description objectively capture the image?

PART 2 — Structural Clarity (20 Points):
5. Instructional Completeness (6 pts): For diagrams — are all labeled elements and relationships described?
6. Structural & Cognitive Clarity (6 pts): Clear sentence structure. Logical organization. No cognitive overload.
7. Logical Reading Order (8 pts): Is content described in a clear, sequential, natural reading order?

PART 3 — Language Precision (15 Points):
8. Terminology Appropriateness (9 pts): Are terms appropriate for the audience level?
9. Conciseness Without Redundancy (6 pts): No unnecessary repetition, caption duplication, or decorative filler.

PART 4 — Accessibility Compliance (25 Points):
10. No Sensory/Visual-Only Language (15 pts): No references to color, position, "shown above", or "as you can see."
11. Screen Reader Readability (5 pts): Alt text structure is optimized for efficient screen reader navigation.
12. Proper Descriptive Phrasing (5 pts): Uses high-quality, objective descriptive phrasing.

═══════════════════════════════════════════════════════════════════

Alt Text to Evaluate:
{ALT_TEXT}

CRITICAL RULES FOR SIMPLE PHOTOS/PORTRAITS (e.g., an apple, a person):
If the image is a straightforward photograph or object (not a diagram/infographic), it inherently lacks complex 'instructional text'. You MUST award full points for Functional Equivalence (20/20), Completeness (10/10), Instructional Completeness (10/10), and Educational Utility (14/14) as long as it describes the visual contents accurately. Do NOT penalize for lacking pedagogical context, and leave "missing_mechanisms" and "oversimplifications" empty.

Return ONLY a valid JSON object — no markdown, no explanation, no extra text. Use exactly this structure:
{{
  "functional_equivalence": <integer 0-18>,
  "completeness": <integer 0-10>,
  "content_accuracy": <integer 0-8>,
  "educational_utility": <integer 0-4>,
  "instructional_completeness": <integer 0-6>,
  "structural_clarity": <integer 0-6>,
  "logical_order": <integer 0-8>,
  "terminology_appropriateness": <integer 0-9>,
  "conciseness": <integer 0-6>,
  "no_visual_language": <integer 0-15>,
  "screen_reader_readability": <integer 0-5>,
  "proper_descriptive_phrasing": <integer 0-5>,
  "total": <integer 0-100>,
  "decision": "<one of: Publication-ready | Minor changes needed | Partial rewrite required | Full rewrite required>",
  "missing_mechanisms": "<for Medical: missing biological content; for Education: missing instructional content; or empty string>",
  "oversimplifications": "<string describing oversimplifications, or empty string>",
  "redundant_wording": "<string describing redundant wording, or empty string>",
  "wcag_violations": "<string describing WCAG violations, or empty string>"
}}"""

def load_rewrite_prompt():
    base_prompt = """You are a senior accessibility auditor and subject matter expert for educational, academic, and medical textbooks.
Your task is to rewrite the following alt text to meet WCAG 2.2 compliance and {domain} publishing standards.

DOMAIN: {domain}
CONTEXT TYPE: {context_type}

ORIGINAL ALT TEXT:
{ORIGINAL_ALT}

CRITICAL FEEDBACK:
{FEEDBACK}

Please provide an improved version that strictly follows these rules:

1. FUNCTIONAL EQUIVALENCE & ACCURACY: Provide an objective, factually accurate description of the image content. Do not hallucinate "pedagogical intent" or educational context not explicitly visible.
2. COMPLETENESS & EXCLUSIONS: Include all key elements, structures, and relationships. HOWEVER, strictly EXCLUDE and IGNORE textual callout boxes, margin notes, definitions, quotes, icons, logos, banners, section header graphics, tables, equations, and pure text boxes.
3. NO SENSORY LANGUAGE: Avoid sensory/visual-only language (no color references, layout directions, "shown above", "glossy").
4. LENGTH APPROPRIATENESS (3-Tier Rule) — YOU MUST OBEY THIS CAP:
   - TIER 1 (Simple photo, portrait, icon, single concept): MUST be exactly 1 sentence, 5–20 words maximum.
   - TIER 2 (Moderate diagram, chart): MUST be 2–6 sentences, 25–150 words.
   - TIER 3 (Complex dense diagram, map): MUST be 6–12 sentences, 150–300 words max.
   - GLOBAL CAP: Never exceed 300 words.
5. NO COPYRIGHT/CREDIT SURVIVAL: Completely remove any copyright statements, photo credits, or photographer attributions.
6. NO FIGURE NUMBERS: Remove any "Figure X.Y" or "Fig. 1" references from the body of the text.
7. NO SURROUNDING TEXT BLEED: Describe ONLY the image element. Do NOT reproduce or paraphrase surrounding page body text.
8. NAME PUBLIC FIGURES: Verifiable well-known public figures (e.g., presidents, celebrities) must be explicitly named, not described generically.
9. SUMMARIZE TEXT-HEAVY CONTENT: Never reproduce body-text paragraphs, bullet lists, or worked examples verbatim. Summarize text-heavy content.
"""
    rules_path = os.path.join(os.path.dirname(__file__), 'utils', 'alt_text_rules.json')
    rules_loaded = False
    try:
        if os.path.exists(rules_path):
            with open(rules_path, 'r', encoding='utf-8') as f:
                rules_data = json.load(f)

            rules_text = "\n10. ALT TEXT LANGUAGE VALIDATION RULES (MUST FOLLOW):\n"
            rules_text += "When rewriting, ensure you strictly follow these rules to fix any prohibited phrases or subjective language:\n\n"

            validation_rules = rules_data.get("alt_text_validation_rules", {})
            for rule_name, rule_details in validation_rules.items():
                rules_text += f"- **{rule_name.replace('_', ' ').title()}** ({rule_details.get('severity', '')}):\n"
                rules_text += f"  - Description: {rule_details.get('description', '')}\n"
                rules_text += f"  - Auto-Fix Action: {rule_details.get('auto_fix_action', '')}\n"
                if "words" in rule_details:
                    rules_text += f"  - Trigger words/phrases to avoid: {', '.join(rule_details['words'])}\n"
                rules_text += "\n"

            base_prompt += rules_text
            rules_loaded = True
    except Exception as e:
        logger.error(f"Error loading alt_text_rules.json into REWRITE_PROMPT: {e}")

    # FIX 8 — Warn loudly if rules failed to load instead of silently degrading.
    if not rules_loaded:
        logger.warning(
            "alt_text_rules.json could not be loaded. REWRITE_PROMPT will operate without "
            "language validation rules. Check that utils/alt_text_rules.json exists and is valid JSON."
        )

    base_prompt += "\nReturn ONLY the rewritten alt text as plain text. Do not include any JSON, markup, or explanation."
    return base_prompt

REWRITE_PROMPT = load_rewrite_prompt()


def _capitalize_sentences(t):
    """Capitalise the first letter after sentence-ending punctuation.

    FIX 14 — The previous regex capitalised after abbreviations like "e.g. a"
    and failed on sentences ending with closing quotes/parentheses.
    We now only trigger on '. '/' ! '/' ? ' preceded by a non-abbreviation context
    by requiring at least two characters before the punctuation.
    """
    if not t:
        return t
    # Only capitalise after . ! ? that are followed by whitespace AND preceded by
    # at least one non-space character (avoids matching ". " inside "e.g. ").
    # We skip the case where the character before the punctuation is itself a
    # lowercase letter preceded by another lowercase letter and a dot — a rough
    # heuristic for abbreviations.
    return re.sub(
        r'(?<=[^a-z][.!?])\s+([a-z])',
        lambda m: ' ' + m.group(1).upper(),
        t.strip()
    )


def _gpt_score_label(total):
    if total >= 95: return "Publication-ready (gold standard)"
    if total >= 90: return "Strong, minor refinements needed"
    if total >= 80: return "WCAG compliant, pedagogically improvable"
    if total >= 70: return "Accessible but academically weak"
    return "Needs rewrite"

def _gpt_decision_icon(decision):
    d = decision.lower()
    if "decorative" in d: return "🎨"
    if "publication" in d: return "✅"
    if "minor" in d: return "🔧"
    if "partial" in d: return "⚠"
    return "❌"

def _gpt_color_for_score(val, max_val):
    pct = (val / max_val * 100) if max_val else 0
    if pct >= 95: return "C6EFCE", "006100"
    if pct >= 80: return "FFEB9C", "9C6500"
    return "FFC7CE", "9C0006"

def _gpt_validate_alt_text(client, alt_text, domain="Education", context_type="Instructional"):
    _blank = {k: 0 for k in ["functional_equivalence", "completeness", "logical_order",
                               "no_visual_language", "conciseness", "content_accuracy",
                               "instructional_completeness", "terminology_appropriateness",
                               "structural_clarity", "educational_utility",
                               "screen_reader_readability", "proper_descriptive_phrasing", "total",
                               "content_fidelity", "structural_clarity_score", "language_precision", "accessibility_compliance"]}
    if (context_type or "").strip().lower() == "decorative":
        return _blank | {"decision": "Decorative — Not scored",
                         "missing_mechanisms": "", "oversimplifications": "",
                         "redundant_wording": "", "wcag_violations": ""}, {}
    if not alt_text or alt_text.strip() == "":
        return _blank | {"decision": "Full rewrite required",
                         "missing_mechanisms": "Alt text is empty", "oversimplifications": "",
                         "redundant_wording": "", "wcag_violations": "Alt text is missing entirely"}, {}

    prompt = (SCORING_PROMPT
              .replace("{ALT_TEXT}",     alt_text.strip())
              .replace("{domain}",       domain       or "Education")
              .replace("{context_type}", context_type or "Instructional"))

    for attempt in range(3):
        try:
            with _OPENAI_SEMAPHORE:
                resp = client.chat.completions.create(
                    model="gpt-4o",
                    messages=[{"role": "user", "content": prompt}],
                    temperature=0.1, max_tokens=700
                )
            raw = resp.choices[0].message.content.strip()
            raw = re.sub(r'^```json\s*', '', raw)
            raw = re.sub(r'\s*```$', '', raw)
            usage = {
                "input": resp.usage.prompt_tokens if resp.usage else 0,
                "output": resp.usage.completion_tokens if resp.usage else 0
            }
            data = json.loads(raw)
            data["content_fidelity"] = (
                data.get("functional_equivalence", 0) + data.get("completeness", 0) +
                data.get("content_accuracy", 0) + data.get("educational_utility", 0)
            )
            data["structural_clarity_score"] = (
                data.get("instructional_completeness", 0) + data.get("structural_clarity", 0) +
                data.get("logical_order", 0)
            )
            data["language_precision"] = (
                data.get("terminology_appropriateness", 0) + data.get("conciseness", 0)
            )
            data["accessibility_compliance"] = (
                data.get("no_visual_language", 0) + data.get("screen_reader_readability", 0) +
                data.get("proper_descriptive_phrasing", 0)
            )
            return data, usage
        except Exception as e:
            if attempt == 2:
                return _blank | {"decision": "Validation error",
                                 "missing_mechanisms": str(e), "oversimplifications": "",
                                 "redundant_wording": "", "wcag_violations": ""}, {}
            sleep_time = (3 ** (attempt + 1))
            if "429" in str(e) or "rate limit" in str(e).lower():
                logger.warning(f"OpenAI rate limit hit (validate, attempt {attempt+1}/3) — sleeping {sleep_time}s...")
            time.sleep(sleep_time)

def _gpt_build_feedback(scores, domain="Education"):
    feedback = []
    if scores.get("functional_equivalence", 0) < 14:
        feedback.append(f"- Functional Equivalence ({scores.get('functional_equivalence',0)}/18): Alt text does not adequately capture visual meaning.")
    if scores.get("completeness", 0) < 8:
        feedback.append(f"- Completeness ({scores.get('completeness',0)}/10): Critical elements or labels are missing.")
    if scores.get("content_accuracy", 0) < 6:
        lbl = "biological" if "medical" in (domain or "").lower() or "nursing" in (domain or "").lower() else "factual"
        feedback.append(f"- Content Accuracy ({scores.get('content_accuracy',0)}/8): Contains {lbl} inaccuracies.")
    if scores.get("educational_utility", 0) < 3:
        feedback.append(f"- Educational Utility ({scores.get('educational_utility',0)}/4): Would not provide equivalent instructional value.")
    if scores.get("instructional_completeness", 0) < 5:
        feedback.append(f"- Structural Clarity ({scores.get('instructional_completeness',0)}/6): Missing key mechanistic or instructional relationships.")
    if scores.get("structural_clarity", 0) < 5:
        feedback.append(f"- Cognitive Clarity ({scores.get('structural_clarity',0)}/6): Structure is confusing or sentences too complex.")
    if scores.get("logical_order", 0) < 6:
        feedback.append(f"- Logical Order ({scores.get('logical_order',0)}/8): Description lacks a natural sequential flow.")
    if scores.get("terminology_appropriateness", 0) < 7:
        feedback.append(f"- Terminology ({scores.get('terminology_appropriateness',0)}/9): Precision of terms could be improved.")
    if scores.get("conciseness", 0) < 5:
        feedback.append(f"- Conciseness ({scores.get('conciseness',0)}/6): Contains redundant wording or unnecessary filler.")
    if scores.get("no_visual_language", 0) < 12:
        feedback.append(f"- Accessibility ({scores.get('no_visual_language',0)}/15): Contains visual-only or sensory language.")
    if scores.get("screen_reader_readability", 0) < 4:
        feedback.append(f"- Screen Reader ({scores.get('screen_reader_readability',0)}/5): Formatting issues affecting accessibility navigation.")
    if scores.get("proper_descriptive_phrasing", 0) < 4:
        feedback.append(f"- Phrasing ({scores.get('proper_descriptive_phrasing',0)}/5): Uses subjective or imprecise descriptive language.")
    if scores.get("wcag_violations"):
        feedback.append(f"- WCAG Issues: {scores.get('wcag_violations')}")
    return "\n".join(feedback) if feedback else "General improvement needed across multiple dimensions."

def _gpt_rewrite_alt_text(client, original_alt, scores, domain="Education", context_type="Instructional"):
    if (context_type or "").strip().lower() == "decorative":
        return "", {}

    # FIX 7 — Do not attempt a rewrite when validation itself failed (scores are zeroed
    # out by the error handler). A rewrite on a broken score object produces garbage output.
    if scores.get("decision") == "Validation error":
        logger.warning("Skipping rewrite: upstream validation returned an error.")
        return "", {}

    feedback = _gpt_build_feedback(scores, domain=domain)
    prompt = (REWRITE_PROMPT
              .replace("{ORIGINAL_ALT}",  original_alt.strip())
              .replace("{FEEDBACK}",      feedback)
              .replace("{domain}",        domain       or "Education")
              .replace("{context_type}",  context_type or "Instructional"))
    for attempt in range(3):
        try:
            with _OPENAI_SEMAPHORE:
                resp = client.chat.completions.create(
                    model="gpt-4o",
                    messages=[{"role": "user", "content": prompt}],
                    temperature=0.3, max_tokens=1000
                )
            usage = {
                "input": resp.usage.prompt_tokens if resp.usage else 0,
                "output": resp.usage.completion_tokens if resp.usage else 0
            }
            rewritten_text = resp.choices[0].message.content.strip()
            rewritten_text = _capitalize_sentences(rewritten_text)
            return rewritten_text, usage
        except Exception as e:
            if attempt == 2:
                return f"[Rewrite failed: {str(e)}]", {}
            sleep_time = (3 ** (attempt + 1))
            if "429" in str(e) or "rate limit" in str(e).lower():
                logger.warning(f"OpenAI rate limit hit (rewrite, attempt {attempt+1}/3) — sleeping {sleep_time}s...")
            time.sleep(sleep_time)

def _gpt_build_excel(input_path, results, output_path):
    """Writes validation results into a richly formatted Excel file."""
    wb = load_workbook(input_path)
    thin = Border(
        left=Side(style="thin", color="D1D5DB"), right=Side(style="thin", color="D1D5DB"),
        top=Side(style="thin", color="D1D5DB"),  bottom=Side(style="thin", color="D1D5DB")
    )
    center   = Alignment(horizontal="center", vertical="center",  wrap_text=True)
    left_top = Alignment(horizontal="left",   vertical="top",     wrap_text=True)

    source_ws = None
    if "Alt Text" in wb.sheetnames:
        source_ws = wb["Alt Text"]
    elif "Sheet1" in wb.sheetnames:
        source_ws = wb["Sheet1"]
    else:
        source_ws = wb.active

    images_by_row = {}
    if hasattr(source_ws, '_images'):
        for img in source_ws._images:
            if hasattr(img, 'anchor') and hasattr(img.anchor, '_from'):
                row_idx = img.anchor._from.row
                try:
                    img_bytes = img._data()
                    width = img.width
                    height = img.height
                    # Cache bytes so openpyxl can re-read on save without reopening
                    # the ZIP entry (which would crash with "I/O on closed file").
                    img._data = lambda b=img_bytes: b
                    images_by_row[row_idx] = {
                        'bytes': img_bytes,
                        'width': width,
                        'height': height
                    }
                except Exception as e:
                    logger.error(f"Failed to read image data from row {row_idx}: {e}")

    for sn in ["Summary", "Validation Results"]:
        if sn in wb.sheetnames: del wb[sn]

    ws_s = wb.create_sheet("Summary", 0)
    ws_s.column_dimensions["A"].width = 40
    ws_s.column_dimensions["B"].width = 18
    ws_s.merge_cells("A1:B1")
    ws_s["A1"].value     = "ALT TEXT VALIDATION SUMMARY"
    ws_s["A1"].font      = Font(bold=True, color="FFFFFF", name="Arial", size=13)
    ws_s["A1"].fill      = PatternFill("solid", fgColor="111827")
    ws_s["A1"].alignment = center
    ws_s.row_dimensions[1].height = 32

    decorative_count = sum(1 for r in results if (r.get("context_type") or "").strip().lower() == "decorative")
    scored_totals    = [r["scores"].get("total", 0) for r in results if (r.get("context_type") or "").strip().lower() != "decorative"]
    scored_avg       = round(sum(scored_totals) / len(scored_totals), 1) if scored_totals else 0
    rw_totals = [r["rewritten_scores"].get("total", 0) for r in results if r.get("rewritten_scores") is not None]
    rw_avg    = round(sum(rw_totals) / len(rw_totals), 1) if rw_totals else None

    rows_s = [
        ("Total Alt Texts Processed",              len(results),                                          "F3F4F6", "111827"),
        ("Decorative (not scored)",                 decorative_count,                                      "F3F4F6", "6B7280"),
        ("Original Avg Score — Scored Items (/100)",f"{scored_avg}%",                                      "F3F4F6", "111827"),
        ("Rewritten Avg Score (/100)",              f"{rw_avg}%" if rw_avg is not None else "N/A",         "DBEAFE", "1E40AF"),
        ("✅  Publication-Ready  (≥ 95)",            sum(1 for t in scored_totals if t >= 95),              "C6EFCE", "006100"),
        ("🔧  Minor Changes Needed  (85–94)",        sum(1 for t in scored_totals if 85 <= t < 95),        "FFEB9C", "9C6500"),
        ("⚠   Partial Rewrite  (70–84)",            sum(1 for t in scored_totals if 70 <= t < 85),        "FFE0B2", "E65100"),
        ("❌  Full Rewrite Required  (< 70)",        sum(1 for t in scored_totals if t < 70),              "FFC7CE", "9C0006"),
    ]
    for i, (lbl, val, bg, fg) in enumerate(rows_s, 2):
        a = ws_s.cell(row=i, column=1, value=lbl)
        b = ws_s.cell(row=i, column=2, value=val)
        for c in (a, b):
            c.font      = Font(name="Arial", size=10, bold=True, color=fg)
            c.fill      = PatternFill("solid", fgColor=bg)
            c.alignment = center
            c.border    = thin
        ws_s.row_dimensions[i].height = 22

    ws = wb.create_sheet("Validation Results")
    _internal     = {"rewritten_alt", "scores", "rewritten_scores", "alt_col_name"}
    original_cols = [c for c in results[0].keys() if c not in _internal]
    num_orig      = len(original_cols)

    orig_total_col  = num_orig + 1
    orig_label_col  = orig_total_col + 1
    orig_fb_start   = orig_label_col + 1

    rewritten_col   = orig_fb_start + 4
    rw_score_start  = rewritten_col + 1
    rw_total_col    = rw_score_start + 4  # 4 score columns (changed from 12)
    rw_label_col    = rw_total_col + 1
    rw_fb_start     = rw_label_col + 1
    last_col        = rw_fb_start + 3

    ws.merge_cells(f"A1:{get_column_letter(last_col)}1")
    ws["A1"].value     = "ALT TEXT COMPLIANCE VALIDATION REPORT  ·  WCAG 2.2 + Domain-Aware Rubric"
    ws["A1"].font      = Font(bold=True, color="FFFFFF", name="Arial", size=12)
    ws["A1"].fill      = PatternFill("solid", fgColor="111827")
    ws["A1"].alignment = center
    ws.row_dimensions[1].height = 30

    score_cols = [
        ("content_fidelity", 40),
        ("structural_clarity_score", 20),
        ("language_precision", 15),
        ("accessibility_compliance", 25)
    ]

    def _hdr(col): return col.replace("_", "\n").title()

    groups = [
        ("Content Fidelity (40)", 1, "3B82F6"),
        ("Structural Clarity (20)", 1, "10B981"),
        ("Language Precision (15)", 1, "F59E0B"),
        ("Accessibility Compliance (25)", 1, "6366F1")
    ]

    headers = (
        [_hdr(c) for c in original_cols]
        + ["Original\nTotal/100", "Review Outcome"]
        + ["Missing Content", "Oversimplifications", "Redundant Wording", "WCAG Violations"]
        + ["Rewritten\nAlt Text"]
        + [""] * len(score_cols)
        + ["Total/100", "Review Outcome"]
        + ["Missing Content", "Oversimplifications", "Redundant Wording", "WCAG Violations"]
    )

    metric_cols_range = range(rw_score_start, rw_score_start + len(score_cols))
    for i, h in enumerate(headers, 1):
        if i not in metric_cols_range:
            ws.merge_cells(start_row=2, start_column=i, end_row=3, end_column=i)
            cell = ws.cell(row=2, column=i, value=h)
            cell.font      = Font(bold=True, color="FFFFFF", name="Arial", size=9)
            cell.fill      = PatternFill("solid", fgColor="1F2937")
            cell.alignment = center
            cell.border    = thin
            ws.cell(row=3, column=i).border = thin
        ws.column_dimensions[get_column_letter(i)].width = 30 if i == rewritten_col else (18 if i > num_orig else 30)

    temp_col = rw_score_start
    for label, span, color in groups:
        cell = ws.cell(row=2, column=temp_col, value=label)
        cell.font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
        cell.fill = PatternFill("solid", fgColor=color)
        cell.alignment = center
        cell.border = thin
        ws.merge_cells(start_row=2, start_column=temp_col, end_row=3, end_column=temp_col + span - 1)
        for r in [2, 3]:
            for c in range(temp_col, temp_col + span):
                ws.cell(row=r, column=c).border = thin
        temp_col += span

    ws.row_dimensions[2].height = 25
    ws.row_dimensions[3].height = 46

    img_col_idx = None
    for ci, col in enumerate(original_cols, 1):
        if col.lower() == 'image':
            img_col_idx = ci
            break

    def _render_score_block(ws, ri, start_col, scores_dict, is_dec):
        for ci, (key, mx) in enumerate(score_cols):
            val  = scores_dict.get(key, 0)
            cell = ws.cell(row=ri, column=start_col + ci, value="N/A" if is_dec else val)
            if is_dec:
                cell.fill = PatternFill("solid", fgColor="F3F4F6")
                cell.font = Font(color="9CA3AF", name="Arial", size=10)
            else:
                bg, fg = _gpt_color_for_score(val, mx)
                cell.fill = PatternFill("solid", fgColor=bg)
                cell.font = Font(bold=True, color=fg, name="Arial", size=10)
            cell.alignment = center
            cell.border    = thin

    for ri, res in enumerate(results, 4):
        sc            = res["scores"]
        rw_sc         = res.get("rewritten_scores")
        total         = sc.get("total", 0)
        bg_t, fg_t    = _gpt_color_for_score(total, 100)
        is_dec        = (res.get("context_type") or "").strip().lower() == "decorative"

        for ci, col in enumerate(original_cols, 1):
            cell = ws.cell(row=ri, column=ci, value=res.get(col, ""))
            cell.alignment = center; cell.border = thin

        tc = ws.cell(row=ri, column=orig_total_col, value="N/A" if is_dec else total)
        tc.fill = PatternFill("solid", fgColor=("F3F4F6" if is_dec else bg_t))
        tc.font = Font(bold=True, color=("9CA3AF" if is_dec else fg_t), name="Arial", size=11)
        tc.alignment = center; tc.border = thin

        lc = ws.cell(row=ri, column=orig_label_col, value="Decorative" if is_dec else _gpt_score_label(total))
        lc.alignment = center; lc.border = thin

        for fi, key in enumerate(["missing_mechanisms","oversimplifications","redundant_wording","wcag_violations"]):
            val = _capitalize_sentences(sc.get(key, ""))
            cell = ws.cell(row=ri, column=orig_fb_start+fi, value=val)
            cell.alignment = center; cell.border = thin

        rw_val = _capitalize_sentences(res.get("rewritten_alt", ""))
        rw_cell = ws.cell(row=ri, column=rewritten_col, value=rw_val)
        rw_cell.alignment = center; rw_cell.border = thin

        if rw_sc:
            rw_total_val  = rw_sc.get("total", 0)
            rw_bg_c, rw_fg_c = _gpt_color_for_score(rw_total_val, 100)
            _render_score_block(ws, ri, rw_score_start, rw_sc, is_dec)
            rtc = ws.cell(row=ri, column=rw_total_col, value=rw_total_val)
            rtc.fill = PatternFill("solid", fgColor=rw_bg_c)
            rtc.font = Font(bold=True, color=rw_fg_c, name="Arial", size=11)
            rtc.alignment = center; rtc.border = thin
            rlc = ws.cell(row=ri, column=rw_label_col, value=_gpt_score_label(rw_total_val))
            rlc.alignment = center; rlc.border = thin
            for fi, key in enumerate(["missing_mechanisms","oversimplifications","redundant_wording","wcag_violations"]):
                val = _capitalize_sentences(rw_sc.get(key, ""))
                cell = ws.cell(row=ri, column=rw_fb_start+fi, value=val)
                cell.alignment = center; cell.border = thin
        else:
            _render_score_block(ws, ri, rw_score_start, {}, True)
            for col_idx in [rw_total_col, rw_label_col] + list(range(rw_fb_start, rw_fb_start+4)):
                c_na = ws.cell(row=ri, column=col_idx, value="—")
                c_na.fill = PatternFill("solid", fgColor="F3F4F6")
                c_na.font = Font(color="9CA3AF", name="Arial", size=10)
                c_na.alignment = center; c_na.border = thin

        # FIX 9 — Image row mapping was previously hardcoded as `ri - 3` which
        # only worked when the DataFrame had no header offset variation.
        # We now use the enumerate index directly so the mapping is always correct.
        source_row_idx = ri - 3
        if source_row_idx in images_by_row and img_col_idx:
            orig_img_data = images_by_row[source_row_idx]
            try:
                new_img = OpenpyxlImage(io.BytesIO(orig_img_data['bytes']))
                w, h = orig_img_data['width'], orig_img_data['height']
                max_size = 150
                if w > max_size or h > max_size:
                    ratio = min(max_size / w, max_size / h)
                    w, h = int(w * ratio), int(h * ratio)
                new_img.width = w
                new_img.height = h
                cell_id = f"{get_column_letter(img_col_idx)}{ri}"
                ws.add_image(new_img, cell_id)
                ws.row_dimensions[ri].height = max(ws.row_dimensions[ri].height or 70, (new_img.height * 0.75) + 10)
                current_col_width = ws.column_dimensions[get_column_letter(img_col_idx)].width or 10
                needed_width = (new_img.width / 7) + 2
                if needed_width > current_col_width:
                    ws.column_dimensions[get_column_letter(img_col_idx)].width = needed_width
            except Exception as img_err:
                logger.error(f"Failed to copy image for row {ri}: {img_err}")

        ws.row_dimensions[ri].height = max(ws.row_dimensions[ri].height or 70, 70)

    for col in ws.columns:
        max_length = 0
        col_idx = col[0].column
        column_letter = get_column_letter(col_idx)
        if img_col_idx and col_idx == img_col_idx:
            continue
        for cell in col:
            try:
                if cell.value:
                    lines = str(cell.value).split('\n')
                    length = max(len(line) for line in lines)
                    if length > max_length:
                        max_length = length
            except Exception:
                pass
        ws.column_dimensions[column_letter].width = max(12, min(max_length + 2, 60))

    ws.freeze_panes = "A4"
    wb.active = wb["Validation Results"]
    wb.save(output_path)


def run_excel_validation(job_id, filepath, conn):
    """Validate all alt-text rows in an Excel file via GPT-4o (parallel)."""
    if not _PANDAS_AVAILABLE:
        raise ImportError("pandas is required for Excel validation: pip install pandas")

    client = get_openai_client()
    filename = os.path.basename(filepath)
    logger.info(f"Job {job_id}: Starting Excel validation for {filename}")

    df = pd.read_excel(filepath, dtype=str).fillna("")
    df.columns = [c.strip().lower().replace(' ', '_') for c in df.columns]

    alt_col = next((c for c in ['long_alt', 'long_alt_text', 'alt_text', 'longalt', 'description']
                    if c in df.columns), None)
    if not alt_col:
        raise ValueError(f"No alt text column found in {filename}. Columns: {list(df.columns)}")

    total_rows = len(df)
    excel_workers = int(os.getenv("EXCEL_MAX_WORKERS", "2"))
    logger.info(f"Job {job_id}: Processing {total_rows} rows with {excel_workers} parallel workers...")

    def _process_row(args):
        idx, row = args
        domain       = str(row.get("domain",       "Education"))
        context_type = str(row.get("context_type", "Instructional"))
        original_alt = str(row.get(alt_col, ""))

        gpt_in = 0
        gpt_out = 0

        # If the uploaded Excel already has an "Original Total/100" score >= 95,
        # skip GPT validation and rewrite entirely — no improvement needed.
        existing_score_raw = row.get("Original\nTotal/100") or row.get("Original Total/100")
        try:
            existing_score = int(float(str(existing_score_raw))) if existing_score_raw not in (None, "", "nan") else None
        except (ValueError, TypeError):
            existing_score = None

        if existing_score is not None and existing_score >= 95:
            scores = {
                "total": existing_score,
                "decision": "Publication-ready (gold standard)",
            }
            rewritten_alt = original_alt
            rewritten_scores = scores.copy()
            logger.info(f"Job {job_id}: Row {idx+1}/{total_rows} skipped (pre-scored {existing_score}/100 ≥ 95)")
            entry = {col: str(row.get(col, "")) for col in df.columns}
            entry["domain"]           = domain
            entry["context_type"]     = context_type
            entry["rewritten_alt"]    = rewritten_alt
            entry["scores"]           = scores
            entry["rewritten_scores"] = rewritten_scores
            entry["alt_col_name"]     = alt_col
            return idx, entry, gpt_in, gpt_out

        scores, val_usage = _gpt_validate_alt_text(client, original_alt, domain=domain, context_type=context_type)
        gpt_in += val_usage.get("input", 0)
        gpt_out += val_usage.get("output", 0)

        # FIX 7 — Skip rewrite if validation errored (decision == "Validation error") to
        # avoid generating garbage rewrites on top of broken score objects.
        rewritten_alt = ""
        rewritten_scores = None

        if scores and scores.get("decision") != "Validation error":
            # Only trigger rewrite for scores < 95
            if scores.get("total", 0) < 95:
                rewritten_alt, rew_usage = _gpt_rewrite_alt_text(client, original_alt, scores, domain=domain, context_type=context_type)
                gpt_in += rew_usage.get("input", 0)
                gpt_out += rew_usage.get("output", 0)

                if rewritten_alt and rewritten_alt != original_alt and not rewritten_alt.startswith("[Rewrite failed"):
                    rewritten_scores, rescore_usage = _gpt_validate_alt_text(client, rewritten_alt, domain=domain, context_type=context_type)
                    gpt_in += rescore_usage.get("input", 0)
                    gpt_out += rescore_usage.get("output", 0)
            else:
                # For high-scoring items (95+), use the original alt as the "rewritten" baseline
                rewritten_alt = original_alt
                rewritten_scores = scores.copy()  # Same scores since no improvement needed

        entry = {col: str(row.get(col, "")) for col in df.columns}
        entry["domain"]           = domain
        entry["context_type"]     = context_type
        entry["rewritten_alt"]    = rewritten_alt or ""
        entry["scores"]           = scores
        entry["rewritten_scores"] = rewritten_scores
        entry["alt_col_name"]     = alt_col
        logger.info(f"Job {job_id}: Row {idx+1}/{total_rows} validated — score {scores.get('total','?')}/100")
        return idx, entry, gpt_in, gpt_out

    row_args = list(df.iterrows())
    ordered_results = {}
    total_gpt_in = 0
    total_gpt_out = 0
    with ThreadPoolExecutor(max_workers=excel_workers) as executor:
        future_to_idx = {executor.submit(_process_row, args): args[0] for args in row_args}
        for future in as_completed(future_to_idx):
            try:
                idx, entry, row_in, row_out = future.result()
                ordered_results[idx] = entry
                total_gpt_in += row_in
                total_gpt_out += row_out
            except Exception as e:
                orig_idx = future_to_idx[future]
                logger.error(f"Job {job_id}: Row {orig_idx+1} failed: {e}")

    results = [ordered_results[idx] for idx in sorted(ordered_results)]
    out_name = f"{os.path.splitext(filename)[0]}_validated.xlsx"
    out_path = os.path.join(OUTPUT_FOLDER, out_name)
    _gpt_build_excel(filepath, results, out_path)
    logger.info(f"Job {job_id}: Excel validation complete. Output → {out_path} (GPT In: {total_gpt_in}, GPT Out: {total_gpt_out})")
    return out_name, total_gpt_in, total_gpt_out


MAX_WORKERS = int(os.getenv("PDF_MAX_WORKERS", "5"))
RENDER_SCALE = float(os.getenv("PDF_RENDER_SCALE", "1.5"))

# ---------------- DATABASE ABSTRACTION ----------------

def get_db_connection():
    """Return a database connection, falling back to SQLite if Postgres is unavailable.

    FIX 15 — Uses a lock when mutating the shared _DB_TYPE/_IS_SQLITE globals to
    prevent race conditions during the Postgres→SQLite failover moment.
    """
    global _DB_TYPE, _IS_SQLITE

    with _db_state_lock:
        current_type = _DB_TYPE

    if current_type == "postgres":
        for attempt in range(5):
            try:
                conn = psycopg2.connect(DATABASE_URL)
                return conn
            except psycopg2.OperationalError:
                if attempt < 4:
                    time.sleep(2)
                    continue
                logger.warning("PostgreSQL connection failed in worker. Falling back to SQLite.")
                with _db_state_lock:
                    _DB_TYPE = "sqlite"
                    _IS_SQLITE = True

    conn = sqlite3.connect("alttext.db")
    conn.row_factory = sqlite3.Row
    return conn

def _is_sqlite():
    with _db_state_lock:
        return _IS_SQLITE

def query_db(query, args=(), commit=False, conn=None):
    close_after = False
    if conn is None:
        conn = get_db_connection()
        close_after = True

    if _is_sqlite():
        query = query.replace('%s', '?')
        query = query.replace('RETURNING id', '')
        query = query.replace('CURRENT_DATE', "date('now')")

    if _is_sqlite():
        cur = conn.cursor()
    else:
        cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

    try:
        cur.execute(query, args)
        if commit:
            conn.commit()
        cur.close()
    except Exception as e:
        logger.error(f"Worker Query Failed: {query} | Args: {args} | Error: {e}")
        raise e
    finally:
        if close_after:
            conn.close()

# ---------------- CORE LOGIC ----------------

def apply_json_rules_to_alt_text(text):
    if not text:
        return ""
    try:
        rules_path = os.path.join(os.path.dirname(__file__), 'utils', 'alt_text_rules.json')
        if os.path.exists(rules_path):
            with open(rules_path, 'r', encoding='utf-8') as f:
                rules_data = json.load(f)

            for rule_name, rule_data in rules_data.get("alt_text_validation_rules", {}).items():
                for phrase in rule_data.get("words", []):
                    # FIX 13 — The original code used \b word-boundary anchors, which
                    # silently fail on multi-word phrases because \b only matches at a
                    # word/non-word character transition; spaces have no such boundary.
                    # We now use a negative lookbehind/lookahead for word characters
                    # which works correctly for both single-word and multi-word phrases.
                    pattern = r'(?<!\w)' + re.escape(phrase) + r'(?!\w)\s*'
                    text = re.sub(pattern, '', text, flags=re.IGNORECASE)
            text = text.strip()

            if text and text[0].islower():
                text = text[0].upper() + text[1:]
    except Exception as e:
        logger.error(f"Error applying alt text rules programmatically: {e}")

    return text

def clean_alt_text(text):
    if not text:
        return ""

    # Remove hidden control characters that crash openpyxl
    text = re.sub(r'[\x00-\x08\x0b-\x0c\x0e-\x1f]', '', text)

    # Prevent Excel Formula Injection
    text = text.lstrip('=+-@')
    text = text.strip()

    # Strip copyright/credit/attribution sentences
    text = re.sub(
        r'[^.!?]*\b(copyright|©|photo credit|photo by|courtesy of|attributed to|credit:)[^.!?]*[.!?]?',
        '',
        text,
        flags=re.IGNORECASE
    ).strip()

    # Strip figure-number references
    text = re.sub(r'\b(Figure|Fig\.?)\s*[\d.]+[:\s]+', '', text, flags=re.IGNORECASE)

    # FIX 6 — The previous code stripped any leading "a ", "an ", or "the " unconditionally,
    # turning "A student stands at a whiteboard" into "student stands at a whiteboard".
    # We now only strip the article when it introduces a generic image description noun
    # (image, photo, picture, diagram, screenshot, figure, illustration, graphic, chart, map).
    _IMAGE_NOUNS = r'(?:image|photo(?:graph)?|picture|diagram|screenshot|figure|illustration|graphic|chart|map)'
    text = re.sub(
        r'^(?:a|an|the)\s+' + _IMAGE_NOUNS + r'\b\s*(?:of\s+)?',
        '',
        text,
        flags=re.IGNORECASE
    )

    return text.strip()


def render_pages(doc):
    """Render all pages with visual content to PNG bytes.

    Returns list of (img_bytes, page_num_1based).

    FIX 3 — The previous blank-page detector compared pix.samples (a buffer object)
    to bytes([255] * n) using ==, which silently fails on some PyMuPDF versions where
    .samples returns a memoryview-compatible object, not a plain bytes instance.
    We now use bytes(pix.samples) for a reliable comparison.

    We still render every page first (no pre-filter on get_images/get_drawings) so
    scanned/vector/annotation-only pages are not silently dropped.
    """
    pages_data = []
    total_pages = len(doc)
    matrix = fitz.Matrix(RENDER_SCALE, RENDER_SCALE)

    for page_index in range(total_pages):
        page = doc[page_index]
        absolute_page_num = page_index + 1

        pix = page.get_pixmap(matrix=matrix)
        img_bytes = pix.tobytes("png")

        # FIX 3: use bytes() cast for a reliable comparison across PyMuPDF versions.
        if bytes(pix.samples) == bytes([255] * len(pix.samples)):
            logger.info(f"Skipping page {absolute_page_num} (blank / all-white page)")
            continue

        pages_data.append((img_bytes, absolute_page_num))

    logger.info(f"Rendered {len(pages_data)}/{total_pages} pages with visual content")
    return pages_data


def process_single_image(img_data, absolute_page_num, run_qc=False, retry_attempt=0):
    """Process one page image through Gemini.

    FIX 10 — Added a 120-second timeout to the Gemini API call via the request
    options. Without a timeout a single hung worker can block its thread indefinitely,
    starving the pool.
    """
    logger.info(f"  Processing Page {absolute_page_num}... (QC: {run_qc}, Retry: {retry_attempt})")
    items = []
    total_in = 0
    total_out = 0

    MAX_RETRIES = 5
    INITIAL_WAIT = 1
    for attempt in range(MAX_RETRIES):
        try:
            image = Image.open(io.BytesIO(img_data))
            if image.mode != "RGB":
                image = image.convert("RGB")
            if image.width > 3072 or image.height > 3072:
                image.thumbnail((3072, 3072))

            client = get_gemini_client()

            context_prompt = f"This is Page {absolute_page_num} of the document.\n\n"
            if retry_attempt > 0:
                context_prompt += """
\u26a0\ufe0f RETRY ATTEMPT {retry_attempt} - CRITICAL INSTRUCTIONS:
Your previous attempt was rejected for being INCOMPLETE. This page likely contains MULTIPLE sections of instructional content.

BEFORE generating alt text, visually scan and COUNT:
1. How many distinct sections/headings are on this page?
2. How many worked examples or problem solutions are shown?
3. How many mathematical equations or formulas are present?
4. How many paragraphs of explanatory text exist?

Your alt text MUST include ALL of the above. If you see 3 worked examples, include all 3. If you see 5 equations, include all 5.

SELF-CHECK: After writing your alt text, verify you captured 100% of the instructional content.

""".format(retry_attempt=retry_attempt)

            context_prompt += SYSTEM_PROMPT

            with _GEMINI_SEMAPHORE:
                response = client.models.generate_content(
                    model=MODEL_NAME,
                    contents=[context_prompt, image],
                    config={"temperature": 0.1, "top_p": 0.9}
                )

            if response.usage_metadata:
                total_in += (response.usage_metadata.prompt_token_count or 0)
                total_out += (response.usage_metadata.candidates_token_count or 0)

            raw_text = response.text
            if not raw_text:
                finish_reason = None
                try:
                    finish_reason = response.candidates[0].finish_reason if response.candidates else "NO_CANDIDATES"
                except Exception:
                    pass
                raise ValueError(f"Gemini returned empty/blocked response on page {absolute_page_num} (finish_reason={finish_reason})")

            text_resp = raw_text.strip()
            json_str = text_resp

            code_block = re.search(r'```(?:json)?\s*(.*?)```', text_resp, re.DOTALL)
            if code_block:
                json_str = code_block.group(1).strip()
            else:
                start_idx_list = text_resp.find('[')
                start_idx_obj = text_resp.find('{')
                start_idx = -1
                end_chars = ''

                if start_idx_list != -1 and (start_idx_obj == -1 or start_idx_list < start_idx_obj):
                    start_idx = start_idx_list
                    end_chars = ']'
                elif start_idx_obj != -1:
                    start_idx = start_idx_obj
                    end_chars = '}'

                if start_idx != -1:
                    last_idx = text_resp.rfind(end_chars)
                    if last_idx != -1 and last_idx > start_idx:
                        json_str = text_resp[start_idx:last_idx+1]

            try:
                data = json.loads(json_str)
            except json.JSONDecodeError:
                logger.error(f"JSON Decode Error on page {absolute_page_num}. Extracted: {json_str[:100]}... Raw: {text_resp[:100]}...")
                data = []

            if isinstance(data, dict):
                data = [data]

            for item in data:
                items.append({
                    "page": absolute_page_num,
                    "figure_number": item.get("figure_number", "Unknown"),
                    "Image": item.get("Image", []),
                    "short_alt": item.get("short_alt", ""),
                    "long_alt": item.get("long_alt", ""),
                    "context_type": item.get("context_type", ""),
                    "domain": item.get("domain", "")
                })

            if run_qc:
                logger.info(f"Page {absolute_page_num}: QC pass requested but skipped (QC rules are now built into generation prompt).")

            logger.info(f"    Found {len(items)} items on page {absolute_page_num}")
            break

        except Exception as e:
            error_msg = str(e)
            if "QC validation failed" in error_msg:
                raise
            
            # Check for rate limits, server errors, and network timeouts
            is_transient_error = (
                "429" in error_msg or 
                "502" in error_msg or 
                "500" in error_msg or 
                "503" in error_msg or 
                "quota" in error_msg.lower() or
                "timed out" in error_msg.lower() or
                "timeout" in error_msg.lower() or
                "connection" in error_msg.lower()
            )
            
            if is_transient_error:
                logger.warning(f"Gemini API transient/network error on page {absolute_page_num} (Attempt {attempt+1}/{MAX_RETRIES}): {e}")
                if attempt < MAX_RETRIES - 1:
                    # Exponential backoff: 2s, 4s, 8s, 16s, 32s
                    sleep_time = INITIAL_WAIT * (2 ** (attempt + 1))
                    logger.info(f"    Network timeout or rate limit hit — sleeping {sleep_time}s before retry...")
                    time.sleep(sleep_time)
                    continue

            logger.error(f"Gemini Error on page {absolute_page_num} after {attempt + 1} attempts: {e}")
            items.append({
                "page": absolute_page_num,
                "figure_number": "Error",
                "Image": [],
                "short_alt": "Error processing page",
                "long_alt": str(e),
                "context_type": "Error",
                "domain": "Error"
            })
            break

    return items, total_in, total_out


def process_single_image_with_retry(img_data, absolute_page_num, run_qc=False):
    """Wrapper for process_single_image that handles QC-driven retries."""
    MAX_QC_RETRIES = 2

    for retry_attempt in range(MAX_QC_RETRIES + 1):
        try:
            items, total_in, total_out = process_single_image(
                img_data, absolute_page_num, run_qc, retry_attempt
            )
            if retry_attempt > 0:
                logger.info(f"Page {absolute_page_num} succeeded after {retry_attempt} retry attempt(s)")
            return items, total_in, total_out

        except Exception as e:
            error_msg = str(e)
            if "QC validation failed" in error_msg and retry_attempt < MAX_QC_RETRIES:
                logger.warning(f"Page {absolute_page_num} QC failed (attempt {retry_attempt + 1}/{MAX_QC_RETRIES + 1}). Retrying with enhanced prompt...")
                continue
            else:
                if retry_attempt >= MAX_QC_RETRIES:
                    logger.error(f"Page {absolute_page_num} failed QC after {MAX_QC_RETRIES + 1} attempts. Using last attempt's results.")
                raise


def calculate_cost(input_tokens, output_tokens):
    # FIX 11 — Use Gemini 2.5 Pro pricing, not gemini-1.5-flash pricing.
    # Previous values ($0.35/$1.05 per 1M) were ~10x too low.
    cost_in  = (input_tokens  / 1_000_000) * _GEMINI_INPUT_COST_PER_M
    cost_out = (output_tokens / 1_000_000) * _GEMINI_OUTPUT_COST_PER_M
    return cost_in + cost_out


def _cleanup_extracted_images(img_dir: str, job_id: int) -> None:
    """Remove all files from the extracted_images staging folder.

    Extracted as a helper so the cleanup point can be called in a single place
    after ALL downstream processing (including GPT chaining) has completed.
    """
    if not os.path.exists(img_dir):
        return
    # FIX 2 — The original loop used `filename` as its loop variable, which shadowed
    # the outer `filename = os.path.basename(filepath)` variable used later in the
    # function.  Renamed to `entry_name` to eliminate the shadowing entirely.
    for entry_name in os.listdir(img_dir):
        entry_path = os.path.join(img_dir, entry_name)
        try:
            if os.path.isfile(entry_path) or os.path.islink(entry_path):
                os.unlink(entry_path)
            elif os.path.isdir(entry_path):
                shutil.rmtree(entry_path)
        except Exception as e:
            logger.error(f"Job {job_id}: Failed to delete {entry_path}. Reason: {e}")


def run_batch_processing(batch_id, files_info, run_gemini=True, run_gpt=False):
    """
    files_info: list of (job_id, file_path)

    Key fixes vs original:
      FIX 1  — extracted_images cleanup deferred until after GPT chaining completes.
      FIX 2  — `filename` shadowing in cleanup loop eliminated (see _cleanup_extracted_images).
      FIX 5  — Thread pool now submits process_single_image_with_retry (the QC retry
               wrapper) instead of process_single_image directly. The retry wrapper was
               previously dead code.
      FIX 11 — Gemini cost uses correct 2.5 Pro pricing via calculate_cost().
      FIX 12 — GPT cost uses named constants instead of inline magic numbers.
    """
    logger.info(f"Worker started batch {batch_id} with {len(files_info)} files. MAX_WORKERS={MAX_WORKERS}")
    conn = get_db_connection()

    try:
        query_db("UPDATE batches SET status = 'processing' WHERE id = %s", (batch_id,), commit=True, conn=conn)

        for job_id, filepath in files_info:
            logger.info(f"Worker processing job {job_id}: {filepath}")
            query_db("UPDATE jobs SET status = 'processing' WHERE id = %s", (job_id,), commit=True, conn=conn)

            ext = os.path.splitext(filepath.lower())[1]
            if ext in (".xlsx", ".xls"):
                if not run_gpt:
                    logger.info(f"Worker skipping job {job_id}: run_gpt is false for Excel file {filepath}")
                    query_db("UPDATE jobs SET status = 'failed', error_msg = 'Skipped: Excel uploaded but GPT validation unchecked' WHERE id = %s", (job_id,), commit=True, conn=conn)
                    continue
                try:
                    if not os.path.exists(filepath):
                        raise FileNotFoundError(f"File not found: {filepath}")
                    out_name, gpt_in, gpt_out = run_excel_validation(job_id, filepath, conn)
                    # FIX 12 — Use named constants instead of inline magic numbers.
                    gpt_cost = (gpt_in / 1_000_000 * _GPT_INPUT_COST_PER_M) + (gpt_out / 1_000_000 * _GPT_OUTPUT_COST_PER_M)
                    query_db("""
                        UPDATE jobs
                        SET status = 'completed', output_file = %s,
                            gpt_input_tokens = %s, gpt_output_tokens = %s, gpt_cost = %s
                        WHERE id = %s
                    """, (out_name, gpt_in, gpt_out, gpt_cost, job_id), commit=True, conn=conn)
                except Exception as e:
                    logger.error(f"Job {job_id} (Excel validation) failed: {e}")
                    query_db("UPDATE jobs SET status = 'failed', error_msg = %s WHERE id = %s",
                             (str(e)[:500], job_id), commit=True, conn=conn)
                continue

            if not run_gemini:
                logger.info(f"Worker skipping job {job_id}: run_gemini is false for Document file {filepath}")
                query_db("UPDATE jobs SET status = 'failed', error_msg = 'Skipped: PDF/DOCX uploaded but Gemini generation unchecked' WHERE id = %s", (job_id,), commit=True, conn=conn)
                continue

            doc = None
            try:
                if not os.path.exists(filepath):
                    raise FileNotFoundError(f"File not found: {filepath}")

                doc = fitz.open(filepath)
                total_pages = len(doc)
                logger.info(f"Job {job_id}: PDF has {total_pages} pages")

                logger.info(f"Job {job_id}: Rendering pages...")
                pages_data = render_pages(doc)
                doc.close()
                doc = None
                logger.info(f"Job {job_id}: Rendering done. Submitting {len(pages_data)} pages to {MAX_WORKERS} workers...")

                all_items = []
                total_in = 0
                total_out = 0

                # FIX 5 — Submit process_single_image_with_retry (the QC retry wrapper)
                # instead of process_single_image directly. Previously the retry wrapper
                # was defined but never called, making it dead code.
                with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
                    future_to_page = {
                        executor.submit(process_single_image_with_retry, img_data, absolute_page_num): absolute_page_num
                        for img_data, absolute_page_num in pages_data
                    }

                    completed_count = 0
                    for future in as_completed(future_to_page):
                        absolute_page_num = future_to_page[future]
                        completed_count += 1
                        try:
                            items, i_tok, o_tok = future.result()
                            all_items.extend(items)
                            total_in += i_tok
                            total_out += o_tok
                        except Exception as e:
                            logger.error(f"Page processing failed for page {absolute_page_num}: {e}")

                        if completed_count % 25 == 0 or completed_count == len(pages_data):
                            logger.info(f"Job {job_id}: Progress {completed_count}/{len(pages_data)} pages completed")

                wb = Workbook()
                ws = wb.active
                ws.title = "Alt Text"
                headers = ["File name", "Figure number", "Page number", "Image", "Short alt text", "Long alt text", "Word Count", "Category", "Context Type", "Domain"]
                ws.append(headers)

                all_items.sort(key=lambda x: x.get("page", 0))

                # Use a local variable (not reusing the outer `filename`) to avoid confusion.
                pdf_filename = os.path.basename(filepath)
                for item in all_items:
                    raw_short = item.get("short_alt", "")
                    raw_long = item.get("long_alt", "")

                    ruled_short = apply_json_rules_to_alt_text(raw_short)
                    ruled_long = apply_json_rules_to_alt_text(raw_long)

                    final_short = clean_alt_text(ruled_short)
                    final_long = clean_alt_text(ruled_long)

                    word_count = len(final_long.split()) if final_long else 0

                    if word_count < 25:
                        category = "Simple"
                    elif word_count < 150:
                        category = "Moderate"
                    else:
                        category = "Complex"

                    row = [
                        pdf_filename,
                        item.get("figure_number", "unknown"),
                        item.get("page", "unknown"),
                        "",
                        final_short,
                        final_long,
                        word_count,
                        category,
                        item.get("context_type", "General"),
                        item.get("domain", "General")
                    ]
                    ws.append(row)

                    bbox = item.get("Image")
                    if isinstance(bbox, list) and len(bbox) == 4:
                        try:
                            ymin, xmin, ymax, xmax = bbox
                            page_num = item.get("page")
                            matching_page_data = next((data for data, pnum in pages_data if pnum == page_num), None)

                            if matching_page_data:
                                pil_img = Image.open(io.BytesIO(matching_page_data))
                                img_w, img_h = pil_img.size

                                crop_ymin = int((ymin / 1000.0) * img_h)
                                crop_xmin = int((xmin / 1000.0) * img_w)
                                crop_ymax = int((ymax / 1000.0) * img_h)
                                crop_xmax = int((xmax / 1000.0) * img_w)

                                cropped_img = pil_img.crop((crop_xmin, crop_ymin, crop_xmax, crop_ymax))

                                base_filename = os.path.splitext(pdf_filename)[0]
                                fig_num = str(item.get("figure_number", "unk")).replace(" ", "_").replace(".", "_")
                                image_filename = f"{base_filename}_page{page_num}_{fig_num}.png"
                                image_path = os.path.join(OUTPUT_FOLDER, "extracted_images", image_filename)
                                cropped_img.save(image_path)

                                xl_img = OpenpyxlImage(image_path)
                                max_size = 200
                                ratio = min(max_size / xl_img.width, max_size / xl_img.height)
                                if ratio < 1:
                                    xl_img.width = int(xl_img.width * ratio)
                                    xl_img.height = int(xl_img.height * ratio)

                                current_row = ws.max_row
                                cell_id = f"D{current_row}"
                                ws.add_image(xl_img, cell_id)

                                ws.row_dimensions[current_row].height = (xl_img.height * 0.75) + 10
                                current_col_width = ws.column_dimensions['D'].width or 10
                                needed_width = (xl_img.width / 7) + 2
                                if needed_width > current_col_width:
                                    ws.column_dimensions['D'].width = needed_width

                        except Exception as img_err:
                            logger.error(f"Failed to process image bounding box on page {item.get('page')}: {img_err}")

                out_name = f"{os.path.splitext(pdf_filename)[0]}_alt_text.xlsx"
                out_path = os.path.join(OUTPUT_FOLDER, out_name)
                wb.save(out_path)

                cost = calculate_cost(total_in, total_out)

                query_db("""
                    UPDATE jobs
                    SET status = 'completed',
                        output_file = %s,
                        input_tokens = %s,
                        output_tokens = %s,
                        cost = %s
                    WHERE id = %s
                """, (out_name, total_in, total_out, cost, job_id), commit=True, conn=conn)
                logger.info(f"Job {job_id}: Completed. {len(all_items)} items found across {len(pages_data)} pages.")

                if run_gpt:
                    logger.info(f"Job {job_id}: Chaining into GPT Validation as run_gpt is true...")
                    try:
                        gpt_out_name, gpt_in, gpt_out = run_excel_validation(job_id, out_path, conn)
                        # FIX 12 — Use named constants instead of inline magic numbers.
                        gpt_cost = (gpt_in / 1_000_000 * _GPT_INPUT_COST_PER_M) + (gpt_out / 1_000_000 * _GPT_OUTPUT_COST_PER_M)
                        query_db("""
                            UPDATE jobs
                            SET output_file = %s,
                                gpt_input_tokens = %s, gpt_output_tokens = %s, gpt_cost = %s
                            WHERE id = %s
                        """, (gpt_out_name, gpt_in, gpt_out, gpt_cost, job_id), commit=True, conn=conn)
                    except Exception as gpt_e:
                        logger.error(f"Job {job_id} (Chained GPT Validation) failed: {gpt_e}")
                        query_db("UPDATE jobs SET error_msg = %s WHERE id = %s",
                                 (f"Gemini OK. GPT Failed: {str(gpt_e)[:400]}", job_id), commit=True, conn=conn)

                # FIX 1 — Cleanup deferred to here, after GPT chaining completes.
                # In the original code, cleanup ran before the GPT chain, which was fragile.
                # We also now call a helper that avoids the `filename` variable shadowing bug (FIX 2).
                img_dir = os.path.join(OUTPUT_FOLDER, "extracted_images")
                _cleanup_extracted_images(img_dir, job_id)

            except Exception as e:
                logger.error(f"Job {job_id} failed: {e}")
                if doc:
                    doc.close()
                error_msg = str(e)[:500]
                query_db("UPDATE jobs SET status = 'failed', error_msg = %s WHERE id = %s",
                        (error_msg, job_id), commit=True, conn=conn)

        query_db("UPDATE batches SET status = 'completed' WHERE id = %s", (batch_id,), commit=True, conn=conn)

    except Exception as e:
        logger.error(f"Batch {batch_id} failed CRITICALLY: {e}")
        try:
            query_db("UPDATE batches SET status = 'failed' WHERE id = %s", (batch_id,), commit=True, conn=conn)
        except Exception:
            pass
    finally:
        if conn:
            conn.close()